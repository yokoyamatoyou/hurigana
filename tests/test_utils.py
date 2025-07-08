import pandas as pd
from unittest.mock import patch
import asyncio

from core.utils import process_dataframe
from core import scorer, utils
from core.utils import to_excel_bytes
from openpyxl import load_workbook, Workbook
from io import BytesIO


def test_process_dataframe_sudachi_match():
    df = pd.DataFrame({
        '名前': ['太郎', '花子'],
        'フリガナ': ['タロウ', 'ハナコ'],
    })

    with patch('core.utils.scorer.gpt_candidates', return_value=[]) as mock:
        out = process_dataframe(df, '名前', 'フリガナ')

    assert list(out['信頼度']) == [95, 95]
    assert list(out['理由']) == ['辞書候補1位一致', '辞書候補1位一致']
    assert mock.call_count == 0


def test_process_dataframe_long_name():
    long_name = 'あ' * 51
    df = pd.DataFrame({'名前': [long_name], 'フリガナ': ['']})

    with patch('core.utils.scorer.gpt_candidates') as mock:
        out = process_dataframe(df, '名前', 'フリガナ')

    assert out['信頼度'][0] == 0
    assert out['理由'][0] == '長すぎる'
    assert mock.call_count == 0


def test_process_dataframe_gpt_called():
    df = pd.DataFrame({'名前': ['未知'], 'フリガナ': ['ミチ']})

    with patch(
        'core.utils.parser.sudachi_reading',
        return_value=None,
    ), patch(
        'core.utils.scorer.gpt_candidates',
        return_value=['ミチ', 'ミチョ'],
    ), patch(
        'core.utils.scorer.calc_confidence',
        wraps=scorer.calc_confidence,
    ) as conf_mock:
        out = process_dataframe(df, '名前', 'フリガナ')

    assert out['信頼度'][0] == 85
    assert out['理由'][0] == '候補1位一致'
    conf_mock.assert_called_once()


def test_process_dataframe_nan_name():
    df = pd.DataFrame({'名前': [pd.NA], 'フリガナ': ['ミチ']})

    with patch('core.utils.scorer.gpt_candidates') as mock:
        out = process_dataframe(df, '名前', 'フリガナ')

    assert out['信頼度'][0] == 0
    assert out['理由'][0] == '長すぎる'
    mock.assert_not_called()


def test_process_dataframe_nan_furigana():
    df = pd.DataFrame({'名前': ['太郎'], 'フリガナ': [pd.NA]})

    with patch('core.utils.scorer.gpt_candidates', return_value=['タロウ']) as mock:
        out = process_dataframe(df, '名前', 'フリガナ')

    assert out['信頼度'][0] == 30
    assert out['理由'][0] == '候補外･要確認'
    mock.assert_called_once_with('太郎')


def test_to_excel_bytes_template():
    df = pd.DataFrame({"A": [1]})
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "old"
    buf = BytesIO()
    wb.save(buf)
    tmpl = buf.getvalue()

    out_bytes = to_excel_bytes(df, template_bytes=tmpl)
    wb2 = load_workbook(BytesIO(out_bytes))
    assert wb2.active["A2"].value == 1


def test_to_excel_bytes_no_template():
    df = pd.DataFrame({"A": [2]})
    out_bytes = to_excel_bytes(df)
    wb = load_workbook(BytesIO(out_bytes))
    assert wb.active["A2"].value == 2


def test_to_excel_bytes_uses_xlsxwriter():
    df = pd.DataFrame({"A": [3]})
    call_engines = []
    orig_writer = pd.ExcelWriter

    def spy_writer(*args, **kwargs):
        call_engines.append(kwargs.get("engine"))
        return orig_writer(*args, **kwargs)

    with patch("pandas.ExcelWriter", side_effect=spy_writer):
        out_bytes = to_excel_bytes(df)

    assert call_engines[0] == "xlsxwriter"
    wb = load_workbook(BytesIO(out_bytes))
    assert wb.active["A2"].value == 3


def test_process_dataframe_progress_and_batch_size():
    df = pd.DataFrame(
        {
            "名前": ["太郎", "花子", "次郎"],
            "フリガナ": ["タロウ", "ハナコ", "ジロウ"],
        }
    )

    progress_calls = []

    def on_progress(done: int, total: int) -> None:
        progress_calls.append((done, total))

    with patch("core.utils.scorer.gpt_candidates", return_value=[]):
        process_dataframe(
            df,
            "名前",
            "フリガナ",
            on_progress=on_progress,
            batch_size=2,
        )

    assert progress_calls == [(1, 3), (2, 3), (3, 3)]


def test_async_process_dataframe_matches_sync():
    df = pd.DataFrame({'名前': ['未知'], 'フリガナ': ['ミチ']})

    async def run_test():
        with patch('core.utils.parser.sudachi_reading', return_value=None), patch(
            'core.utils.scorer.async_gpt_candidates', return_value=['ミチ']
        ) as g_mock:
            out = await utils.async_process_dataframe(df, '名前', 'フリガナ', batch_size=1)
        assert g_mock.called
        return out

    result = asyncio.run(run_test())

    with patch('core.utils.parser.sudachi_reading', return_value=None), patch(
        'core.utils.scorer.gpt_candidates', return_value=['ミチ']
    ):
        expected = process_dataframe(df, '名前', 'フリガナ')
    pd.testing.assert_frame_equal(result, expected)


def test_async_process_dataframe_deduplicates():
    df = pd.DataFrame({'名前': ['未知', '未知'], 'フリガナ': ['ミチ', 'ミチ']})

    async def run_test():
        with patch('core.utils.parser.sudachi_reading', return_value=None), patch(
            'core.utils.scorer.async_gpt_candidates', return_value=['ミチ']
        ) as g_mock:
            out = await utils.async_process_dataframe(df, '名前', 'フリガナ', batch_size=2)
        assert g_mock.call_count == 1
        return out

    result = asyncio.run(run_test())
    assert list(result['信頼度']) == [85, 85]
    assert list(result['理由']) == ['候補1位一致', '候補1位一致']


def test_process_dataframe_deduplicates():
    df = pd.DataFrame({'名前': ['未知', '未知'], 'フリガナ': ['ミチ', 'ミチ']})

    with patch('core.utils.parser.sudachi_reading', return_value=None), patch(
        'core.utils.scorer.gpt_candidates', return_value=['ミチ']
    ) as g_mock:
        out = process_dataframe(df, '名前', 'フリガナ', batch_size=1)

    assert g_mock.call_count == 1
    assert list(out['信頼度']) == [85, 85]
    assert list(out['理由']) == ['候補1位一致', '候補1位一致']
